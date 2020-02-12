from openpyxl import load_workbook

import pandas as pd

for i in range(0, 10):
    data_xls = pd.read_excel('/Users/sun/Documents/2nd_round_result_{}.xlsx'.format(i), 'Sheet1', index_col=None)
    data_xls.to_csv('2nd_result.csv', encoding='utf-8', index=False, mode='a', sep='\t')
# wb = load_workbook(filename='/Users/sun/Downloads/exp_result.xlsx')
# sheet = wb.active
#
# results_wb = load_workbook(filename='/Users/sun/Downloads/result_0.xlsx')
# result_sheet = wb.active
# print(list(result_sheet.rows))
# for row in result_sheet.rows:
#     value_list = []
#     print(row)

# for cell in row:
#     value_list.append(cell.value)
# sheet.append(value_list)
# print(value_list)
# value_list.clear()
# results_wb = load_workbook(filename='/Users/sun/Downloads/result_1.xlsx')
# result_sheet = wb.active
# sheet.append(list(result_sheet.rows)[1:])

# wb.save('/Users/sun/Downloads/exp_result.xlsx')
